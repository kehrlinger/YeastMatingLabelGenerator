# This script takes two vector and their respective inserts as arguments and
# combines them in the same way a mating assay would be carried out.
# An excel spreadsheet is created and the generated labels are saved to this spreadsheet
import openpyxl
import string
import datetime
import os.path

vec1 = input('Please insert the first vector: \n(for example: pGADT7)\n')
ins_vec_1 = input ('Please enter the inserts of this vector, delimited by a whitespace: \n'
                           '(for example: insert1 insert2 insert3)\n')

vec2 = input('Please enter the second vector: \n'
             '(for example: pGBKT7)\n')
ins_vec_2 = input('Please enter the inserts of the second vector, delimited by a whitespace: \n'
                  '(for example: insert1 insert2 insert3)\n')

def construct_vectors(vector, inserts):
    vector_list = []
    for i in inserts:
        i = vector + '-' + i
        vector_list.append(i)
    return vector_list

final_ins_vec_1 = construct_vectors(vec1, ins_vec_1.split())
final_ins_vec_2 = construct_vectors(vec2, ins_vec_2.split())

labels = []
for v in final_ins_vec_1:
    for i in range(0, len(final_ins_vec_2)):
        label = v + ' x ' + final_ins_vec_2[i]
        labels.append(label)

# Construct a list with coordinates
coordinates =[]
letters = list(string.ascii_uppercase)
letter_index = 0
loopcount_coordinates = 0
for c in range(0, len(labels)):
    #print (loopcount_coordinates)
    if loopcount_coordinates < 15:
        loopcount_coordinates += 1
        coordinate = letters[letter_index] + str(loopcount_coordinates)
        coordinates.append(coordinate)
    else:
        loopcount_coordinates = 1
        letter_index += 1
        coordinate = letters[letter_index] + str(loopcount_coordinates)
        coordinates.append(coordinate)

# Write everything to a workbook
# Check if the workbook TubeLabels.xlsx already exists, proceed accordingly
# WB is saved to Tubelabels.xlsx, if it exists a new sheet is generated
# Sheetname is 'Date Labels'
if os.path.isfile('./TubeLabels.xlsx'):
    wb = openpyxl.load_workbook('TubeLabels.xlsx')
    timestamp = datetime.date.today()
    sheet = wb.create_sheet(index=0, title= str(timestamp) + ' Labels')
    for index in range(0, len(labels)):
        sheet[coordinates[index]] = labels[index]
    wb.save('TubeLabels.xlsx')
else:
    timestamp = datetime.date.today()
    # Create a new workbook. I still don't fully understand openpyxl
    wb = openpyxl.workbook.Workbook()
    sheet = wb.active
    sheet.title = str(timestamp) + ' Labels'
    for index in range(0, len(labels)):
        sheet[coordinates[index]] = labels[index]
    wb.save('TubeLabels.xlsx')
